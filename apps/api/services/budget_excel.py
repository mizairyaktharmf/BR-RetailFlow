"""
Budget Excel Parser — Smart Header Detection
Reads DAILY SALES TRACKER .xlsx files of ANY column order/format.
Scans all header rows and matches columns by their actual label text.
"""

import io
import logging
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return default


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    try:
        # Try parsing common date strings
        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(str(val).strip(), fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue
    except Exception:
        pass
    return str(val)


def _normalize(text: str) -> str:
    """Normalize label for matching — uppercase, strip spaces/special chars."""
    return str(text).upper().replace(" ", "").replace(".", "").replace("-", "").replace("_", "").replace("/", "")


# ── Column matcher rules ────────────────────────────────────────────────────────
# Each rule: (field_name, list_of_keywords_that_must_ALL_be_in_normalized_label)
# Earlier rules take priority. We do two passes: one for section labels (MTD vs Day)

# Fields that belong to the MTD (right-side) section
MTD_FIELDS = [
    ("mtd_budget",   ["MTD", "BUDGET"]),
    ("mtd_cur",      ["MTD", "2026"]),
    ("mtd_ly",       ["MTD", "2025"]),
    ("mtd_grth",     ["MTD", "GRTH"]),
    ("mtd_grth",     ["MTD", "GROWTH"]),
    ("mtd_ach",      ["MTD", "ACH"]),
]

# Fields that belong to the Day Sales / Guest Count (left-side) section
DAY_FIELDS = [
    ("sl",           ["SL"]),
    ("day",          ["DAY"]),
    ("date",         ["DATE"]),
    ("date",         ["2026"]),          # "2026" date column
    ("ly_sales",     ["2025"]),          # "2025" = last year sales (before guest count section)
    ("cur_sales",    ["2026"]),          # second 2026 col = current sales
    ("budget",       ["BUDGET"]),
    ("grth_sales",   ["GRTH"]),
    ("grth_sales",   ["GROWTH"]),
    ("ach_sales",    ["ACH"]),
    ("ly_gc",        ["GC", "2025"]),    # GC + 2025 = LY guest count
    ("ly_gc",        ["GUEST", "2025"]),
    ("cur_gc",       ["GC", "2026"]),
    ("cur_gc",       ["GUEST", "2026"]),
    ("grth_gc",      ["GC", "GRTH"]),
    ("grth_gc",      ["GC", "GROWTH"]),
]


def _scan_headers(ws) -> dict:
    """
    Scan ALL rows in the first 8 rows for header labels.
    Build a merged label per column by concatenating non-empty header text
    (handles merged cells / multi-row headers like 'Days Sales' over '2025 | 2026').

    Returns dict: field_name -> column_index (1-based)
    """
    max_col = ws.max_column

    # Collect all text per column across header rows
    col_labels = {}   # col_index -> list of normalized text fragments
    for row in range(1, 9):
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip():
                norm = _normalize(str(val))
                col_labels.setdefault(col, []).append(norm)

    # For each column, build a single combined label string
    col_combined = {col: "".join(frags) for col, frags in col_labels.items()}
    logger.debug(f"Column combined labels: {col_combined}")

    # Identify which columns are "MTD section" by looking for MTD anywhere in header stack
    mtd_cols = {col for col, text in col_combined.items() if "MTD" in text}

    assigned = {}  # field -> col

    # Pass 1: assign MTD fields
    for field, keywords in MTD_FIELDS:
        if field in assigned:
            continue
        for col, text in col_combined.items():
            if col not in mtd_cols:
                continue
            if all(kw in text for kw in keywords):
                assigned[field] = col
                logger.info(f"  Detected MTD field '{field}' -> col {col} (label: {text[:60]})")
                break

    # Pass 2: assign Day fields (non-MTD columns only)
    # Track which columns are already assigned to avoid double-use
    used_cols = set(assigned.values())
    # Keep separate counters for fields that can appear multiple times
    date_assigned = False
    sales_2025_assigned = False
    sales_2026_assigned = False
    gc_2025_assigned = False
    gc_2026_assigned = False

    for field, keywords in DAY_FIELDS:
        if field in assigned:
            continue
        for col in sorted(col_combined.keys()):  # left to right
            if col in mtd_cols:
                continue
            if col in used_cols and field not in ("date", "ly_sales", "cur_sales", "ly_gc", "cur_gc"):
                continue
            text = col_combined.get(col, "")
            if not all(kw in text for kw in keywords):
                continue

            # Special disambiguation for repeated column patterns
            if field == "date" and date_assigned:
                continue
            if field == "ly_sales" and sales_2025_assigned:
                continue
            if field == "cur_sales" and sales_2026_assigned:
                continue
            if field == "ly_gc" and gc_2025_assigned:
                continue
            if field == "cur_gc" and gc_2026_assigned:
                continue

            assigned[field] = col
            used_cols.add(col)
            logger.info(f"  Detected Day field '{field}' -> col {col} (label: {text[:60]})")

            if field == "date":         date_assigned = True
            if field == "ly_sales":     sales_2025_assigned = True
            if field == "cur_sales":    sales_2026_assigned = True
            if field == "ly_gc":        gc_2025_assigned = True
            if field == "cur_gc":       gc_2026_assigned = True
            break

    logger.info(f"Final column map: {assigned}")
    return assigned


def _fallback_columns() -> dict:
    """
    Hard-coded fallback matching the Daily Sales Tracker screenshot:
    A=SL B=Date C=DAY D=LY2025 E=2026 F=Grth% G=Budget H=Ach%
    I=LY_GC J=2026_GC K=Grth% L=MTD_LY M=MTD_2026 N=MTD_Grth% O=MTD_Budget P=MTD_Ach%
    """
    return {
        "sl": 1, "date": 2, "day": 3,
        "ly_sales": 4, "cur_sales": 5, "grth_sales": 6,
        "budget": 7, "ach_sales": 8,
        "ly_gc": 9, "cur_gc": 10, "grth_gc": 11,
        "mtd_ly": 12, "mtd_cur": 13, "mtd_grth": 14,
        "mtd_budget": 15, "mtd_ach": 16,
    }


# ── KPI scanner ────────────────────────────────────────────────────────────────

def _scan_kpis(ws, start_row: int) -> dict:
    """
    After the TOTAL row, scan remaining rows for KPI table.
    Looks for labels like ATV, AUV, Cake QTY, HP QTY in any column,
    then reads the numeric value from adjacent columns.
    """
    kpis = {}
    for row in range(start_row + 1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 6) + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            label = _normalize(str(val))
            # Find the first numeric value to the right of this label
            num_val = None
            for nc in range(col + 1, col + 5):
                nv = ws.cell(row=row, column=nc).value
                f = _safe_float(nv, default=None)
                if f is not None and f != 0:
                    num_val = f
                    break

            if "ATV" in label and "CAUV" not in label and "atv" not in kpis:
                kpis["atv"] = num_val or _safe_float(ws.cell(row=row, column=col + 1).value)
            elif "AUV" in label and "auv" not in kpis:
                kpis["auv"] = num_val or _safe_float(ws.cell(row=row, column=col + 1).value)
            elif "CAKE" in label and "cake_qty" not in kpis:
                kpis["cake_qty"] = num_val
            elif label.startswith("HP") and "hp_qty" not in kpis:
                kpis["hp_qty"] = num_val

    return kpis


# ── Header row scanner ─────────────────────────────────────────────────────────

def _find_header_info(ws) -> dict:
    """
    Scan top rows for parlor name, month, area manager.
    Looks for keywords like 'PARLOR', 'STORE', 'MONTH', 'AREA MANAGER'.
    """
    info = {"parlor_name": None, "month": None, "area_manager": None}
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            label = _normalize(str(val))
            # Check adjacent cell for value
            next_val = ws.cell(row=row, column=col + 1).value
            if "PARLOR" in label or "STORE" in label or "BRANCHNAME" in label or "ORNAME" in label:
                info["parlor_name"] = str(next_val).strip() if next_val else None
            elif label in ("MONTH", "MONTH:"):
                info["month"] = str(next_val).strip() if next_val else None
            elif "AREAMANAGER" in label or "AREAMGR" in label:
                info["area_manager"] = str(next_val).strip() if next_val else None
    return info


# ── Main parser ────────────────────────────────────────────────────────────────

def parse_budget_excel(file_bytes: bytes) -> dict:
    """
    Parse a DAILY SALES TRACKER Excel file of ANY column format.

    Strategy:
    1. Scan header rows → auto-detect column positions by label text
    2. Fall back to hardcoded layout if detection misses critical fields
    3. Parse daily rows by SL number (1-31)
    4. Find TOTAL row by scanning for "TOTAL" text
    5. Scan below TOTAL for KPI table (ATV, AUV, Cake QTY, HP QTY)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Step 1: detect headers
    cols = _scan_headers(ws)

    # Step 2: fill missing critical fields from fallback
    fallback = _fallback_columns()
    for key, fb_col in fallback.items():
        if key not in cols:
            cols[key] = fb_col
            logger.info(f"  Fallback: '{key}' -> col {fb_col}")

    # Step 3: scan for parlor/month/manager info
    header_info = _find_header_info(ws)

    def get(r, key):
        return ws.cell(row=r, column=cols[key]).value

    daily_data = []
    totals_row = None

    # Step 4: scan data rows
    for row in range(1, ws.max_row + 1):
        # Check for TOTAL row (any of first 4 cells contains "TOTAL")
        row_snippet = " ".join(
            str(ws.cell(row=row, column=c).value or "") for c in range(1, 5)
        ).upper()
        if "TOTAL" in row_snippet and totals_row is None:
            totals_row = row
            continue

        if totals_row is not None:
            break  # stop parsing data rows after TOTAL

        # Must have a numeric SL (1-31)
        sl_val = ws.cell(row=row, column=cols["sl"]).value
        try:
            sl = int(float(sl_val))
            if sl < 1 or sl > 31:
                continue
        except (ValueError, TypeError):
            continue

        ly_sales   = _safe_float(get(row, "ly_sales"))
        cur_sales  = _safe_float(get(row, "cur_sales"))
        budget     = _safe_float(get(row, "budget"))
        ly_gc      = _safe_int(get(row, "ly_gc"))
        cur_gc     = _safe_int(get(row, "cur_gc"))
        mtd_ly     = _safe_float(get(row, "mtd_ly"))
        mtd_cur    = _safe_float(get(row, "mtd_cur"))
        mtd_budget = _safe_float(get(row, "mtd_budget"))

        day_entry = {
            "sl": sl,
            "date_2026": _parse_date(get(row, "date")),
            "date_2025": None,
            "day": str(get(row, "day") or ""),
            "days_sales": {
                "ly_2025":         ly_sales  or None,
                "current_2026":    cur_sales or None,
                "growth_pct":      _safe_float(get(row, "grth_sales")) or None,
                "budget":          budget    or None,
                "achievement_pct": _safe_float(get(row, "ach_sales")) or None,
            },
            "days_guest_count": {
                "ly_2025":      ly_gc  or None,
                "current_2026": cur_gc or None,
                "growth_pct":   _safe_float(get(row, "grth_gc")) or None,
            },
            "mtd_sales": {
                "ly_2025":         mtd_ly     or None,
                "current_2026":    mtd_cur    or None,
                "growth_pct":      _safe_float(get(row, "mtd_grth")) or None,
                "budget":          mtd_budget or None,
                "achievement_pct": _safe_float(get(row, "mtd_ach")) or None,
            },
        }
        daily_data.append(day_entry)

    # Step 5: scan KPIs below TOTAL
    kpi_rows = _scan_kpis(ws, totals_row or ws.max_row) if totals_row else {}

    # Step 6: totals row values
    def totals_cell(key):
        return ws.cell(row=totals_row, column=cols[key]).value if totals_row else None

    totals = {
        "ly_sales_total":      _safe_float(totals_cell("ly_sales")),
        "budget_total":        _safe_float(totals_cell("budget")),
        "ly_gc_total":         _safe_int(totals_cell("ly_gc")),
        "current_sales_total": _safe_float(totals_cell("cur_sales")) or None,
        "current_gc_total":    _safe_int(totals_cell("cur_gc")) or None,
    }

    wb.close()

    # Step 7: build KPIs structure
    kpis = {
        "atv":      {"ly_2025": kpi_rows.get("atv"),      "current_2026": None, "diff_vs_py": None},
        "auv":      {"ly_2025": kpi_rows.get("auv"),      "current_2026": None, "diff_vs_py": None},
        "cake_qty": {"ly_2025": kpi_rows.get("cake_qty"), "current_2026": None, "diff_vs_py": None},
        "hp_qty":   {"ly_2025": kpi_rows.get("hp_qty"),   "current_2026": None, "diff_vs_py": None},
    }

    # Step 8: calculate derived fields per day
    ly_atv_overall = kpi_rows.get("atv") or 0
    for day in daily_data:
        ly_gc_d   = (day["days_guest_count"]["ly_2025"])  or 0
        ly_sal_d  = (day["days_sales"]["ly_2025"])        or 0
        budget_d  = (day["days_sales"]["budget"])         or 0

        day_ly_atv = (ly_sal_d / ly_gc_d) if ly_gc_d > 0 else ly_atv_overall
        day["_ly_atv"] = round(day_ly_atv, 2)
        day["_budget_gc"] = round(budget_d / day_ly_atv) if budget_d > 0 and day_ly_atv > 0 else 0
        day["_budget_atv"] = round(budget_d / day["_budget_gc"], 2) if day["_budget_gc"] > 0 else 0

    logger.info(
        f"Parsed {len(daily_data)} days, "
        f"total_budget={totals['budget_total']}, "
        f"ly_sales={totals['ly_sales_total']}, "
        f"parlor={header_info.get('parlor_name')}"
    )

    return {
        "header": {
            "title": "DAILY SALES TRACKER",
            "parlor_name": header_info.get("parlor_name"),
            "month": header_info.get("month"),
            "month_code": None,
            "area_manager": header_info.get("area_manager"),
        },
        "daily_data": daily_data,
        "totals": totals,
        "kpis": kpis,
    }
